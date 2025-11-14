#-*-coding:utf-8-*-#
import sys

import os
import codecs
import json
import types
import hashlib
from openpyxl import load_workbook

config_path = "config.json"
md5_path = "config.md5"

#特殊语义
teshuyuu = '"'

def is_havs_file(file_name):
	return os.path.exists(file_name)
	
def loal_file(file_name):
	file = codecs.open(file_name, 'r', 'utf-8')
	text = file.read()
	file.close()
	return text

def save_lua(data, path):
	file = codecs.open(path, 'w', 'utf-8')
	file.write(data)
	file.close()


if __name__ == '__main__':
	md5_json_str = "{}"

	md5_path_root = os.path.expanduser('~') + "/" + md5_path
	isHave = is_havs_file(md5_path_root)

	if isHave:
		#加载md5记录
		md5_json_str = loal_file(md5_path_root)

	config_md5_json = json.loads(md5_json_str)

	#解析config json
	config_json = json.loads(loal_file(os.path.abspath(config_path)))

	#自动加载lua
	auto_requrice_text = "" 

	#自动分配后端加载任务
	# auto_server_new = "local tConfigMap = {\n"

	for item_json in config_json["array"]:
		#读取到配置信息
		excel_name_path = os.path.abspath(item_json["excel_name"])

		output_path = os.path.abspath("output/" + item_json["output_path"])
		sheet_name = item_json["sheet_name"]

		is_array = item_json["is_array"]

		#自动写入
		output_file_name = output_path.split("\\")[-1]
		output_file_name = output_file_name.split(".")[0]

		# duixiangname = "cf_" + output_file_name

		# auto_requrice_text = auto_requrice_text + duixiangname + " = require \"config/" + item_json["output_path"].split(".")[0] + "\"  \n"
		# auto_server_new = auto_server_new + output_file_name + " = " + duixiangname + ",\n"

		curmd5 = hashlib.md5(open(excel_name_path,'rb').read()).hexdigest()

		# if output_file_name in config_md5_json:
		# 	if curmd5 == config_md5_json[output_file_name]:
		# 		continue

		config_md5_json[output_file_name] = curmd5

		print("----------build lua ->", output_file_name)
		
		wb = load_workbook(excel_name_path, data_only=True)
		ws = wb.get_sheet_by_name(sheet_name)

		lua_content = "{ \"root\":["

		for row_index in range(4, ws.max_row+1):

			#获取类型
			s_format = str(ws.cell(row = 1, column = 1).value)
	
			strkey = str(ws.cell(row = row_index, column = 1).value)
	
			if strkey == "None":
				continue
			
			print("----{:.2f}%".format(row_index/(ws.max_row+1)*100))
			
			lua_content = lua_content + "{"		
	
			for columns_index in range(1, ws.max_column+1):
	
				s_format = str(ws.cell(row = 1, column = columns_index).value)
	
				if s_format == "None":
					#清除上一个逗号
					if lua_content.endswith(","):
						lua_content = lua_content.rstrip(",")
					continue
	
				if str(ws.cell(row = 2, column = columns_index).value) == "None":
					continue
					
				if s_format == "str":
					#key
					lua_content = lua_content + "\"" + str(ws.cell(row = 2, column = columns_index).value) + "\":"
					#value
					strvalue = str(ws.cell(row = row_index, column = columns_index).value)
					if strvalue == "None":
						strvalue = ""
	
					strvalue = strvalue.replace(teshuyuu, "\\\"")
	
					strvalue = strvalue.replace("\n", "\\n")
					
					lua_content = lua_content + "\"" + strvalue + "\""
				else:
					#key
					lua_content = lua_content + "\"" + str(ws.cell(row = 2, column = columns_index).value) + "\":"
					#value
					strvalue = str(ws.cell(row = row_index, column = columns_index).value)
	
					if strvalue == "None" or strvalue == " " or strvalue == "":
						strvalue = "0"
	
					lua_content = lua_content + strvalue
	
				if columns_index != ws.max_column:
					lua_content = lua_content + ","
			
			if row_index == ws.max_row :
				lua_content = lua_content + "}"
			else:
				nextkey = str(ws.cell(row = row_index+1, column = 1).value)
				if nextkey != "None":
					lua_content = lua_content + "},"
				else:
					lua_content = lua_content + "}"
		lua_content = lua_content + "\n] } \n"
		
		save_lua(lua_content, output_path)

	save_lua(json.dumps(config_md5_json), md5_path_root)

